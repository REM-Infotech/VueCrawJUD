"""Module for creating XLSX templates from list templates."""

import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from pytz import timezone
from quart import current_app as app


class MakeModels:
    """Class to generate XLSX files based on a model and header lists."""

    def __init__(self, model_name: str, display_name: str) -> None:
        """Initialize a MakeModels instance.

        Args:
            model_name (str): The model or key identifier.
            display_name (str): The display name for the file.

        """
        self.model_name = model_name
        self.displayname = display_name

    def make_output(self) -> tuple[str, str]:
        """Create an XLSX file with headers based on the model list.

        Returns:
            tuple: A tuple with the file path and the file name.

        """
        dir_file = Path(__file__).parent.resolve().joinpath("models")

        temp_dir = app.config["TEMP_DIR"]
        os.makedirs(temp_dir, exist_ok=True)
        name_file = f"{self.displayname.upper()} - {datetime.now(timezone('Etc/GMT+4')).strftime('%H-%M-%S')}.xlsx"
        itens_append_file = dir_file.joinpath(f"{self.model_name}.json")
        path_template = str(Path(temp_dir).joinpath(name_file))

        # Criar um novo workbook e uma planilha
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Resultados", 0)
        sheet = workbook.active

        # Cabeçalhos iniciais
        cabecalhos = ["NUMERO_PROCESSO"]
        list_to_append = []

        if not itens_append_file.exists():
            self.model_name = self.model_name.split("_")[-1]
            itens_append_file = dir_file.joinpath(f"{self.model_name}.json")
            if not itens_append_file.exists():
                self.model_name = "without_model.json"
                itens_append_file = dir_file.joinpath(f"{self.model_name}.json")

        with itens_append_file.open("r") as f:
            itens_append = json.loads(f.read())
            list_to_append.extend(itens_append)

        cabecalhos.extend(list_to_append)
        # Definir estilo
        my_red = openpyxl.styles.colors.Color(rgb="A6A6A6")
        my_fill = PatternFill(patternType="solid", fgColor=my_red)
        bold_font = Font(name="Times New Roman", italic=True)

        # Escrever os cabeçalhos na primeira linha da planilha e aplicar estilo
        for pos, coluna in enumerate(cabecalhos):
            item = sheet.cell(row=1, column=pos + 1, value=coluna.upper())
            item.font = bold_font
            item.fill = my_fill

        # Ajustar a largura das colunas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if cell.value:
                    max_length = max(len(str(cell.value)), max_length)
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(path_template)
        return path_template, name_file
