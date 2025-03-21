"""Provide functionality to create and customize Excel files.

Using openpyxl, build an Excel file with headers and styles based on CrawJUD attributes.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

from crawjud.bot.core import CrawJUD
from crawjud.bot.utils.MakeTemplate.appends import Listas

# from typing import list


class MakeXlsx(CrawJUD):
    """Create and customize an Excel file with configured headers and styles.

    Inherits:
        CrawJUD: Base class with common CrawJUD functionalities.
    """

    def __init__(self) -> None:
        """Initialize the MakeXlsx instance.

        No additional parameters are required during initialization.
        """

    def make_output(self, type_xlsx: str, path_template: str) -> list[str]:
        """Build and save an Excel file with customized headers and styles.

        Args:
            type_xlsx (str): String defining Excel template type.
            path_template (str): File system path to save the generated Excel file.

        Returns:
            list[str]: The list of headers used in the created Excel file.

        """
        lista_colunas: list[str] = getattr(Listas(), f"{self.typebot}_{type_xlsx}", getattr(Listas(), type_xlsx, None))
        # Create workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Resultados", 0)
        sheet = workbook.active

        # Initial headers
        cabecalhos = ["NUMERO_PROCESSO"]
        list_to_append = []

        list_to_append.extend(lista_colunas)
        cabecalhos.extend(list_to_append)

        # Define style settings
        my_red = openpyxl.styles.colors.Color(rgb="A6A6A6")
        my_fill = PatternFill(patternType="solid", fgColor=my_red)
        bold_font = Font(name="Times New Roman", italic=True)

        # Write headers in first row with styles
        for pos, coluna in enumerate(cabecalhos):
            item = sheet.cell(row=1, column=pos + 1, value=coluna.upper())
            item.font = bold_font
            item.fill = my_fill

        # Adjust column widths based on header length
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get column name
            for cell in col:
                try:
                    max_length = max(len(str(cell.value)), max_length)
                except Exception as e:
                    raise e
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # Save the workbook at the specified path
        workbook.save(path_template)

        return cabecalhos
